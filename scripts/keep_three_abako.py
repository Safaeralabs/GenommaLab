"""Utility to trim Accesob2b.xlsx to three Abako providers."""

from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    root = Path(__file__).resolve().parent.parent
    workbook_path = root / "Accesob2b.xlsx"
    if not workbook_path.exists():
        raise SystemExit(f"{workbook_path} no existe.")

    workbook = load_workbook(workbook_path)
    sheet = workbook["01_Accesos_Estructurados"]
    header_row = 4
    rows_to_keep: list[tuple[object, ...]] = []

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        portal_value = (row[4] or "").strip().lower()
        if portal_value == "abako":
            rows_to_keep.append(row)
        if len(rows_to_keep) == 3:
            break

    if len(rows_to_keep) < 3:
        raise SystemExit("No se encontraron al menos 3 filas de Abako.")

    sheet.delete_rows(header_row + 1, sheet.max_row - header_row)
    for row_values in rows_to_keep:
        sheet.append(row_values)

    tmp_path = workbook_path.with_suffix(".tmp.xlsx")
    if tmp_path.exists():
        tmp_path.unlink()
    workbook.save(tmp_path)
    tmp_path.replace(workbook_path)
    workbook.close()
    print("Accesob2b.xlsx ahora contiene solo tres filas de Abako.")


if __name__ == "__main__":
    main()
