"""Consolidate sales/inventory rows into the shared Homologaciones workbook."""

from __future__ import annotations

import csv
import logging
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.config import settings
from app.core.models import HomologationSummary, OrganizedFile, Proveedor, ProviderRunDetail


HOMOLOGATION_COLUMNS = [
    "año",
    "Semana",
    "Tipo",
    "Fecha_Stock",
    "Cadena",
    "Cod_Prod",
    "Descripcion_prod",
    "Cod_Local",
    "Descripcion_Local",
    "Unidades",
    "Zonalocal",
]

# Column index (1-based) for the "Cadena" column in the data sheet.
_CADENA_COL_IDX = 5  # col E


@dataclass(slots=True)
class HomologationRow:
    tipo: str
    cod_prod: str
    descripcion: str
    unidades: str
    cadena: str
    cod_local: str
    descripcion_local: str
    zonalocal: str


class HomologationWriter:
    """Build a consolidated Homologaciones file from template + collected rows."""

    SALES_MAPPING = {
        "codigoarticulo": "Cod_Prod",
        "codigoproducto": "Cod_Prod",
        "codprod": "Cod_Prod",
        "referencia": "Cod_Prod",             # Xeon Paretto CSV (Pastor Julio)
        "descripcionarticulo": "Descripcion_prod",
        "descripcionproducto": "Descripcion_prod",
        "descripcion": "Descripcion_prod",
        "unidades": "Unidades",
        "cantidad": "Unidades",
        "venta": "Unidades",                  # Xeon Paretto CSV: columna Venta
        "ventas": "Unidades",                 # Xeon Paretto CSV: columna Ventas
        "ean": "Cod_Prod",
        "nombrepr": "Descripcion_prod",       # NOMBRE_PR normalized
        "nombreproducto": "Descripcion_prod", # NOMBRE_PRODUCTO (EOS)
        "ventasunidad": "Unidades",           # ventas unidad normalized
        "ventaunidad": "Unidades",            # alternate
        "ventaunidades": "Unidades",          # VENTA_UNIDADES (EOS)
    }

    INVENTORY_MAPPING = {
        "codigoarticulo": "Cod_Prod",
        "codigoproducto": "Cod_Prod",
        "articulo": "Descripcion_prod",
        "producto": "Descripcion_prod",
        "disponible": "Unidades",
        "existencia": "Unidades",
        # Aliases EOS Consultores (inventario derivado del delta de ventas)
        "nombreproducto": "Descripcion_prod",
        "ventaunidades": "Unidades",
        "ventaunidad": "Unidades",
        # Xeon / Pastor Julio (Lista de Precios)
        "referencia": "Cod_Prod",
        "descripcion": "Descripcion_prod",
        "stockreal": "Unidades",
        "stockfisico": "Unidades",
        "stock": "Unidades",
    }

    def __init__(self, logger: logging.Logger) -> None:
        self.logger = logger
        self.template_path = settings.TEMPLATE_HOMOLOGACION

    def collect_rows(
        self,
        proveedor: Proveedor,
        organized_files: list[OrganizedFile],
    ) -> list[HomologationRow]:
        rows: list[HomologationRow] = []
        if not organized_files:
            self.logger.warning("[%s] collect_rows: sin archivos organizados.", proveedor.display_name)
            return rows

        for item in organized_files:
            mapping = (
                self.SALES_MAPPING if item.category == "ventas" else self.INVENTORY_MAPPING
            )
            tipo = "SO" if item.category == "ventas" else "INV"
            file_size_kb = item.path.stat().st_size / 1024 if item.path.exists() else -1
            self.logger.info(
                "[%s] Extrayendo %s de '%s' (%.1f KB, ext=%s).",
                proveedor.display_name, tipo, item.path.name,
                file_size_kb, item.path.suffix.lower(),
            )
            entries = self._extract_entries(item.path, mapping)
            if not entries:
                self.logger.warning(
                    "[%s] %s: 0 filas extraidas de '%s'. "
                    "Verifica que las columnas del archivo coincidan con el mapping.",
                    proveedor.display_name, tipo, item.path.name,
                )
            else:
                self.logger.info(
                    "[%s] %s: %d filas extraidas de '%s'.",
                    proveedor.display_name, tipo, len(entries), item.path.name,
                )
            for cod, desc, unidades in entries:
                rows.append(
                    HomologationRow(
                        tipo=tipo,
                        cod_prod=cod,
                        descripcion=desc,
                        unidades=unidades,
                        cadena=proveedor.cadena or proveedor.proveedor,
                        cod_local="",
                        descripcion_local=proveedor.sede_subportal or proveedor.display_name,
                        zonalocal=proveedor.sede_subportal or "",
                    )
                )

        total_so = sum(1 for r in rows if r.tipo == "SO")
        total_inv = sum(1 for r in rows if r.tipo == "INV")
        self.logger.info(
            "[%s] collect_rows total: %d filas (SO=%d, INV=%d).",
            proveedor.display_name, len(rows), total_so, total_inv,
        )
        return rows

    def write(
        self,
        rows: list[HomologationRow],
        year: int,
        week: int,
        fecha_stock: str,
        total_providers: int = 0,
        missing_providers: list[str] | None = None,
        provider_details: list[ProviderRunDetail] | None = None,
    ) -> HomologationSummary:
        if missing_providers is None:
            missing_providers = []

        filename = f"Homologaciones_S{week:02d}_{year}.xlsx"
        target_path = settings.POSTPROCESSED_DIR / filename

        # Build a set of (cadena, tipo) pairs being updated so we only replace
        # the specific types provided — partial re-runs don't erase the other type.
        updating_cadena_tipos: set[tuple[str, str]] = {(row.cadena, row.tipo) for row in rows}
        updating_cadenas = {row.cadena for row in rows}

        if target_path.exists():
            workbook = load_workbook(target_path)
            # Get the data sheet (prefer "homologacion", fall back to active)
            if "homologacion" in workbook.sheetnames:
                sheet = workbook["homologacion"]
            else:
                sheet = workbook.active

            # Keep rows whose cadena is not being updated, OR whose cadena is
            # being updated but its tipo is NOT included in the new data.
            # This allows partial re-runs (e.g. only ventas) to preserve
            # previously loaded data of the other type (e.g. inventario).
            kept_rows: list[tuple] = []
            for row_values in sheet.iter_rows(min_row=2, values_only=True):
                if not row_values or all(v is None for v in row_values):
                    continue
                existing_cadena = row_values[_CADENA_COL_IDX - 1]
                existing_tipo = row_values[2]  # Tipo is column C (index 2, 0-based)
                if (existing_cadena, existing_tipo) not in updating_cadena_tipos:
                    kept_rows.append(row_values)

            # Clear all rows after the header, then rewrite kept + new rows
            max_row = sheet.max_row
            if max_row > 1:
                sheet.delete_rows(2, max_row - 1)

            for kept in kept_rows:
                sheet.append(list(kept))
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "homologacion"
            sheet.append(HOMOLOGATION_COLUMNS)

        # Write new rows
        for row in rows:
            sheet.append(
                [
                    str(year),
                    str(week),
                    row.tipo,
                    fecha_stock,
                    row.cadena,
                    row.cod_prod,
                    row.descripcion,
                    row.cod_local,
                    row.descripcion_local,
                    row.unidades,
                    row.zonalocal,
                ]
            )

        # ── Build "Resumen" sheet ─────────────────────────────────────────────
        # Read all rows now in the data sheet to build the summary
        all_rows_in_sheet: list[tuple] = []
        for row_values in sheet.iter_rows(min_row=2, values_only=True):
            if not row_values or all(v is None for v in row_values):
                continue
            all_rows_in_sheet.append(row_values)

        # Gather per-cadena info and row counts from data sheet
        cadena_tipos: dict[str, set[str]] = {}
        cadena_tipo_count: dict[tuple[str, str], int] = defaultdict(int)
        for rv in all_rows_in_sheet:
            cadena_val = rv[_CADENA_COL_IDX - 1]
            tipo_val = rv[2]
            if cadena_val is None:
                continue
            cadena_key = str(cadena_val)
            tipo_key = str(tipo_val) if tipo_val else ""
            cadena_tipos.setdefault(cadena_key, set()).add(tipo_key)
            cadena_tipo_count[(cadena_key, tipo_key)] += 1

        # Build lookup cadena → ProviderRunDetail from current run
        detail_by_cadena: dict[str, ProviderRunDetail] = {}
        if provider_details:
            for det in provider_details:
                detail_by_cadena[det.cadena] = det

        # Delete existing Resumen sheet if present
        if "Resumen" in workbook.sheetnames:
            del workbook["Resumen"]

        resumen = workbook.create_sheet("Resumen")
        resumen_header = ["Proveedor", "Cadena", "Ventas (SO)", "Filas SO", "Inventario (INV)", "Filas INV", "Estado", "Detalle"]
        resumen.append(resumen_header)
        for cell in resumen[1]:
            cell.font = Font(bold=True)

        # Color fills por estado
        fill_completo  = PatternFill("solid", fgColor="C6EFCE")
        fill_parcial   = PatternFill("solid", fgColor="FFEB9C")
        fill_faltante  = PatternFill("solid", fgColor="FFCC99")
        fill_error     = PatternFill("solid", fgColor="FFC7CE")
        fill_sin_datos = PatternFill("solid", fgColor="EDEDED")

        estado_fill = {
            "Completo":  fill_completo,
            "Parcial":   fill_parcial,
            "Faltante":  fill_faltante,
            "Error":     fill_error,
            "Sin datos": fill_sin_datos,
        }

        def _append_resumen(row_data: list, estado: str) -> None:
            resumen.append(row_data)
            fill = estado_fill.get(estado)
            if fill:
                for cell in resumen[resumen.max_row]:
                    cell.fill = fill

        # 1. Portales con datos en el fichero
        cadenas_procesadas: set[str] = set()
        for cadena_key, tipos in sorted(cadena_tipos.items()):
            cadenas_procesadas.add(cadena_key)
            has_so = "SO" in tipos
            has_inv = "INV" in tipos
            so_count = cadena_tipo_count.get((cadena_key, "SO"), 0)
            inv_count = cadena_tipo_count.get((cadena_key, "INV"), 0)

            det = detail_by_cadena.get(cadena_key)
            display = det.display_name if det else cadena_key

            if has_so and has_inv:
                estado = "Completo"
            elif has_so or has_inv:
                estado = "Parcial"
            else:
                estado = "Faltante"

            if det and not det.success:
                detalle = det.message[:300]
            elif det and det.success:
                detalle = "OK"
            else:
                detalle = "Ejecución previa"

            _append_resumen([
                display, cadena_key,
                "✓" if has_so else "—", so_count if has_so else "—",
                "✓" if has_inv else "—", inv_count if has_inv else "—",
                estado, detalle,
            ], estado)

        # 2. Portales del run actual sin datos (fallidos o sin filas extraíbles)
        if provider_details:
            for det in provider_details:
                if det.cadena in cadenas_procesadas:
                    continue
                estado = "Sin datos" if det.success else "Error"
                _append_resumen([
                    det.display_name, det.cadena,
                    "—", "—", "—", "—",
                    estado, det.message[:300],
                ], estado)

        # Ajustar anchos de columnas del Resumen
        col_widths = [30, 20, 12, 10, 16, 12, 12, 60]
        for i, width in enumerate(col_widths, start=1):
            resumen.column_dimensions[resumen.cell(1, i).column_letter].width = width

        settings.POSTPROCESSED_DIR.mkdir(parents=True, exist_ok=True)
        workbook.save(target_path)
        workbook.close()

        all_cadenas = set(cadena_tipos.keys())
        included = len(all_cadenas)

        # En re-runs parciales, total_providers viene del run actual (ej. 1 portal).
        # Usamos el maximo entre el total declarado y los proveedores reales en el fichero,
        # para que el UI refleje siempre el estado global acumulado.
        effective_total = max(total_providers, included)

        # Los missing son los que NO tienen datos en el fichero final (ni SO ni INV).
        # Combinamos los missing del run actual con los que ya estaban ausentes antes.
        current_missing_in_file = [
            name for name in missing_providers if name not in all_cadenas
        ]

        self.logger.info(
            "Homologación acumulativa S%s/%s: %s (filas nuevas: %s, proveedores incluidos: %s/%s)",
            week,
            year,
            target_path,
            len(rows),
            included,
            effective_total,
        )

        return HomologationSummary(
            path=target_path,
            included_providers=included,
            total_providers=effective_total,
            missing_providers=current_missing_in_file,
        )

    def _ensure_template(self) -> None:
        if self.template_path.exists():
            return
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "homologacion"
        sheet.append(HOMOLOGATION_COLUMNS)
        workbook.save(self.template_path)
        workbook.close()

    def _ensure_header(self, sheet) -> None:
        header = [cell.value for cell in sheet[1]]
        if header and all(h in header for h in HOMOLOGATION_COLUMNS):
            return
        sheet.insert_rows(1)
        for col_idx, column in enumerate(HOMOLOGATION_COLUMNS, start=1):
            sheet.cell(row=1, column=col_idx, value=column)

    def _extract_entries(
        self,
        file_path: Path,
        mapping: dict[str, str],
    ) -> list[tuple[str, str, str]]:
        if not file_path.exists():
            self.logger.error("Archivo no encontrado: %s", file_path)
            return []
        ext = file_path.suffix.lower()
        if ext == ".csv":
            return self._extract_entries_csv(file_path, mapping)
        if ext == ".xls":
            return self._extract_entries_xls(file_path, mapping)
        if ext == ".xlsx":
            return self._extract_entries_xlsx(file_path, mapping)
        self.logger.warning("Extension no soportada para homologacion: '%s' (%s)", ext, file_path.name)
        return []

    def _extract_entries_xlsx(
        self,
        file_path: Path,
        mapping: dict[str, str],
    ) -> list[tuple[str, str, str]]:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        header_row_index, headers = self._find_header_row(sheet, mapping)
        if header_row_index is None:
            # Loguear las primeras filas para diagnóstico
            preview = []
            for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
                preview.append([str(c) for c in row if c is not None])
            self.logger.warning(
                "XLSX '%s': no se encontro fila de cabecera con los campos requeridos. "
                "Primeras filas: %s",
                file_path.name, preview,
            )
            workbook.close()
            return []

        lookup: dict[str, int] = {}
        for idx, raw_label in enumerate(headers):
            normalized_label = self._normalize_text(raw_label)
            target = mapping.get(normalized_label)
            if target:
                lookup[target] = idx

        required = {"Cod_Prod", "Descripcion_prod", "Unidades"}
        missing = required - lookup.keys()
        if missing:
            raw_cols = [str(h) for h in headers if h]
            self.logger.warning(
                "XLSX '%s': faltan campos requeridos %s. "
                "Columnas detectadas (raw): %s | Columnas mapeadas: %s",
                file_path.name, missing, raw_cols, list(lookup.keys()),
            )
            workbook.close()
            return []

        self.logger.debug(
            "XLSX '%s': cabecera en fila %d, mapeado: %s",
            file_path.name, header_row_index, lookup,
        )
        entries: list[tuple[str, str, str]] = []
        for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            if not row:
                continue
            cod = row[lookup["Cod_Prod"]]
            desc = row[lookup["Descripcion_prod"]]
            unidades = row[lookup["Unidades"]]
            if not cod or not desc:
                continue
            entries.append((str(cod).strip(), str(desc).strip(), str(unidades or "").strip()))

        workbook.close()
        return entries

    def _extract_entries_csv(
        self,
        file_path: Path,
        mapping: dict[str, str],
    ) -> list[tuple[str, str, str]]:
        """Read a CSV file and extract homologation entries.

        Dots in numeric fields (decimal/thousands separators) are replaced
        with commas to match the expected homologation format.
        """
        used_encoding = None
        all_rows: list[list[str]] = []
        for encoding in ("utf-8-sig", "latin-1", "utf-8"):
            try:
                with open(file_path, newline="", encoding=encoding) as f:
                    sample = f.read(4096)
                    f.seek(0)
                    try:
                        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                        delimiter = getattr(dialect, "delimiter", ",")
                    except csv.Error:
                        dialect = csv.excel  # type: ignore[assignment]
                        delimiter = ","
                    reader = csv.reader(f, dialect)
                    all_rows = list(reader)
                used_encoding = encoding
                break
            except UnicodeDecodeError:
                continue

        if not all_rows:
            self.logger.warning("CSV '%s': archivo vacio o no decodificable.", file_path.name)
            return []

        self.logger.debug(
            "CSV '%s': %d filas totales, encoding=%s, delimitador='%s'.",
            file_path.name, len(all_rows), used_encoding, delimiter,
        )

        # Find header row (same logic as xlsx: first row matching required targets)
        required = {"Cod_Prod", "Descripcion_prod", "Unidades"}
        header_idx = None
        lookup: dict[str, int] = {}
        for row_idx, row in enumerate(all_rows[:10]):
            found: set[str] = set()
            candidate_lookup: dict[str, int] = {}
            for col_idx, raw_label in enumerate(row):
                norm = self._normalize_text(str(raw_label))
                target = mapping.get(norm)
                if target:
                    found.add(target)
                    candidate_lookup[target] = col_idx
            if required.issubset(found):
                header_idx = row_idx
                lookup = candidate_lookup
                break

        if header_idx is None:
            # Loguear las primeras filas reales para saber qué columnas tiene el archivo
            preview_cols = [all_rows[i] for i in range(min(3, len(all_rows)))]
            self.logger.warning(
                "CSV '%s': no se encontro cabecera con los campos requeridos %s. "
                "Primeras filas del archivo: %s",
                file_path.name, required, preview_cols,
            )
            return []

        raw_header = all_rows[header_idx]
        self.logger.debug(
            "CSV '%s': cabecera en fila %d -> %s | mapeado: %s",
            file_path.name, header_idx, raw_header, lookup,
        )

        entries: list[tuple[str, str, str]] = []
        for row in all_rows[header_idx + 1:]:
            if not row:
                continue
            try:
                cod = row[lookup["Cod_Prod"]].strip()
                desc = row[lookup["Descripcion_prod"]].strip()
                unidades_raw = row[lookup["Unidades"]].strip()
                # Reemplazar puntos por comas en el campo numérico de unidades
                unidades = unidades_raw.replace(".", ",")
            except IndexError:
                continue
            if not cod or not desc:
                continue
            entries.append((cod, desc, unidades))

        return entries

    def _find_header_row(
        self,
        sheet,
        mapping: dict[str, str],
    ) -> tuple[int | None, list[str]]:
        """Locate the first row that can serve as the header for the mapping."""
        required_targets = {"Cod_Prod", "Descripcion_prod", "Unidades"}
        for row_index in range(1, min(10, sheet.max_row) + 1):
            row = [cell.value or "" for cell in sheet[row_index]]
            found_targets: set[str] = set()
            for raw_label in row:
                normalized_label = self._normalize_text(str(raw_label))
                target = mapping.get(normalized_label)
                if target:
                    found_targets.add(target)
            if required_targets.issubset(found_targets):
                return row_index, [str(cell) for cell in row]
        return None, []

    def _extract_entries_xls(
        self,
        file_path: Path,
        mapping: dict[str, str],
    ) -> list[tuple[str, str, str]]:
        """Read a legacy .xls file using xlrd and extract homologation entries."""
        import xlrd  # type: ignore[import]

        try:
            wb = xlrd.open_workbook(str(file_path))
        except Exception:
            return []

        sheet = wb.sheet_by_index(0)
        required = {"Cod_Prod", "Descripcion_prod", "Unidades"}

        header_idx = None
        lookup: dict[str, int] = {}
        for row_idx in range(min(10, sheet.nrows)):
            row = [str(sheet.cell_value(row_idx, c)) for c in range(sheet.ncols)]
            found: set[str] = set()
            candidate: dict[str, int] = {}
            for col_idx, raw_label in enumerate(row):
                norm = self._normalize_text(raw_label)
                target = mapping.get(norm)
                if target:
                    found.add(target)
                    candidate[target] = col_idx
            if required.issubset(found):
                header_idx = row_idx
                lookup = candidate
                break

        if header_idx is None:
            # Fallback posicional para archivos sin cabecera estándar (ej. Provecol)
            return self._extract_entries_xls_positional(sheet)

        entries: list[tuple[str, str, str]] = []
        for row_idx in range(header_idx + 1, sheet.nrows):
            try:
                cod = str(sheet.cell_value(row_idx, lookup["Cod_Prod"])).strip()
                desc = str(sheet.cell_value(row_idx, lookup["Descripcion_prod"])).strip()
                raw_u = sheet.cell_value(row_idx, lookup["Unidades"])
                unidades = str(raw_u).strip().replace(".", ",")
            except IndexError:
                continue
            if not cod or not desc:
                continue
            entries.append((cod, desc, unidades))

        return entries

    @staticmethod
    def _extract_entries_xls_positional(
        sheet,
    ) -> list[tuple[str, str, str]]:
        """Extracción posicional para XLS sin cabecera estándar (formato Provecol).

        Detecta automáticamente si es ventas (código en col 3, desc en col 7)
        o inventario (código en col 1, desc en col 3) buscando la primera fila
        con un código numérico de 6 dígitos.
        """
        import re
        CODE_RE = re.compile(r"^\d{3,}[A-Za-z]{0,3}$")

        # Detectar formato buscando la columna donde aparecen los códigos
        col_code, col_desc, col_units = None, None, None
        for row_idx in range(sheet.nrows):
            row = [str(sheet.cell_value(row_idx, c)).strip() for c in range(sheet.ncols)]
            # Formato ventas: código en col 3, descripción en col 7
            if (len(row) > 14 and CODE_RE.match(row[3])
                    and row[7] and not CODE_RE.match(row[7])):
                col_code, col_desc, col_units = 3, 7, 14
                break
            # Formato inventario: código en col 1, descripción en col 3
            if (len(row) > 18 and CODE_RE.match(row[1])
                    and row[3] and not CODE_RE.match(row[3])):
                col_code, col_desc, col_units = 1, 3, 18
                break

        if col_code is None:
            return []

        entries: list[tuple[str, str, str]] = []
        for row_idx in range(sheet.nrows):
            row = [str(sheet.cell_value(row_idx, c)).strip() for c in range(sheet.ncols)]
            if len(row) <= col_units:
                continue
            cod = row[col_code]
            desc = row[col_desc]
            raw_u = row[col_units]
            if not CODE_RE.match(cod) or not desc:
                continue
            unidades = raw_u.replace(".", ",")
            entries.append((cod, desc, unidades))

        return entries

    @staticmethod
    def _normalize_text(value: str) -> str:
        normalized = unicodedata.normalize("NFKD", value.strip().lower())
        return "".join(ch for ch in normalized if ch.isalnum())
