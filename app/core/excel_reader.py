"""Excel loading logic for provider definitions."""

from __future__ import annotations

import unicodedata
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from app.config import settings
from app.core.models import Proveedor


EXPECTED_COLUMNS = [
    "proveedor",
    "activo",
    "portal_tipo",
    "login_url",
    "usuario",
    "password",
    "fecha_desde",
    "fecha_hasta",
    "carpeta",
]

ACCESOB2B_SHEET_NAME = "01_Accesos_Estructurados"
ACCESOB2B_HEADER_ROW = 4
ACCESOB2B_COLUMNS = [
    "Estado",
    "Activo_RPA",
    "Requiere_revision",
    "Fuente",
    "Portal",
    "ID_original",
    "Cliente",
    "Canal_grupo",
    "Sede_subportal",
    "URL_principal",
    "URL_alternativa",
    "Usuario",
    "Password",
    "Tipo_acceso",
    "Notas_operativas",
    "Conflictos_detectados",
]

CANONICAL_PORTAL_MAP = {
    "abako": "abako",
    "portal_a": "abako",
    "eosconsultores": "eos_consultores",
    "solucionespracticas": "soluciones_practicas",
    "xeon": "xeon",
    "portal_b": "portal_b",
}


class ExcelReader:
    """Read provider configuration from an Excel file."""

    def __init__(self, excel_path: Path) -> None:
        self.excel_path = excel_path

    def read_proveedores(self) -> list[Proveedor]:
        """Load active providers from the configured workbook."""
        if not self.excel_path.exists():
            raise FileNotFoundError(f"No se encontro el Excel: {self.excel_path}")

        workbook = load_workbook(self.excel_path, data_only=True)

        if settings.EXCEL_SHEET_NAME in workbook.sheetnames:
            return self._read_standard_sheet(workbook[settings.EXCEL_SHEET_NAME])

        if ACCESOB2B_SHEET_NAME in workbook.sheetnames:
            return self._read_accesob2b_sheet(workbook[ACCESOB2B_SHEET_NAME])

        raise ValueError(
            "No se encontro una hoja compatible. "
            f"Esperadas: '{settings.EXCEL_SHEET_NAME}' o '{ACCESOB2B_SHEET_NAME}'."
        )

    def _read_standard_sheet(self, sheet) -> list[Proveedor]:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [self._as_text(value) for value in rows[0]]
        if headers != EXPECTED_COLUMNS:
            raise ValueError(
                "Las columnas del Excel no coinciden con el formato esperado. "
                f"Esperadas: {EXPECTED_COLUMNS}"
            )

        proveedores: list[Proveedor] = []
        for row_index, row in enumerate(rows[1:], start=2):
            if self._is_empty_row(row):
                continue

            data = dict(zip(headers, row))
            proveedor = Proveedor(
                proveedor=self._as_text(data["proveedor"]),
                activo=self._as_bool(data["activo"]),
                portal_tipo=self._normalize_portal_type(data["portal_tipo"]),
                portal_origen=self._as_text(data["portal_tipo"]),
                login_url=self._as_text(data["login_url"]),
                usuario=self._as_text(data["usuario"]),
                password=self._as_text(data["password"]),
                fecha_desde=self._as_date_text(data["fecha_desde"]),
                fecha_hasta=self._as_date_text(data["fecha_hasta"]),
                carpeta=self._as_text(data["carpeta"]),
            )

            if not proveedor.proveedor:
                raise ValueError(f"La fila {row_index} no tiene valor en 'proveedor'.")

            if proveedor.activo:
                proveedores.append(proveedor)

        return proveedores

    def _read_accesob2b_sheet(self, sheet) -> list[Proveedor]:
        header_values = next(
            sheet.iter_rows(
                min_row=ACCESOB2B_HEADER_ROW,
                max_row=ACCESOB2B_HEADER_ROW,
                values_only=True,
            )
        )
        headers = [self._as_text(value) for value in header_values]
        if headers != ACCESOB2B_COLUMNS:
            raise ValueError(
                "La hoja estructurada de Accesob2b no tiene las columnas esperadas. "
                f"Esperadas: {ACCESOB2B_COLUMNS}"
            )

        proveedores: list[Proveedor] = []
        for row_index, row in enumerate(
            sheet.iter_rows(min_row=ACCESOB2B_HEADER_ROW + 1, values_only=True),
            start=ACCESOB2B_HEADER_ROW + 1,
        ):
            if self._is_empty_row(row):
                continue

            data = dict(zip(headers, row))
            proveedor_nombre = self._as_text(data["Cliente"])
            sede_subportal = self._as_text(data["Sede_subportal"])
            proveedor = Proveedor(
                proveedor=proveedor_nombre,
                activo=self._as_bool(data["Activo_RPA"]),
                portal_tipo=self._normalize_portal_type(data["Portal"]),
                portal_origen=self._as_text(data["Portal"]),
                login_url=self._as_text(data["URL_principal"])
                or self._as_text(data["URL_alternativa"]),
                usuario=self._as_text(data["Usuario"]),
                password=self._as_text(data["Password"]),
                fecha_desde="",
                fecha_hasta="",
                carpeta=self._build_folder_name(proveedor_nombre, sede_subportal),
                sede_subportal=sede_subportal,
                requiere_revision=self._as_bool(data["Requiere_revision"]),
                notas_operativas=self._as_text(data["Notas_operativas"]),
                conflictos_detectados=self._as_text(data["Conflictos_detectados"]),
                fuente=self._as_text(data["Fuente"]),
                tipo_acceso=self._as_text(data["Tipo_acceso"]),
            )

            if not proveedor.proveedor:
                raise ValueError(f"La fila {row_index} no tiene valor en 'Cliente'.")

            if proveedor.activo:
                proveedores.append(proveedor)

        return proveedores

    @staticmethod
    def _is_empty_row(row: tuple[object, ...]) -> bool:
        return all(value in (None, "") for value in row)

    @staticmethod
    def _as_text(value: object) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @classmethod
    def _as_bool(cls, value: object) -> bool:
        if isinstance(value, bool):
            return value
        if value is None:
            return False
        text = cls._normalize_text(value)
        return text in {"1", "true", "si", "yes", "x", "activo"}

    @staticmethod
    def _as_date_text(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, date):
            return value.isoformat()
        return str(value).strip()

    @staticmethod
    def _normalize_text(value: object) -> str:
        text = str(value).strip().lower()
        return "".join(
            char
            for char in unicodedata.normalize("NFKD", text)
            if not unicodedata.combining(char)
        )

    @classmethod
    def _normalize_portal_type(cls, value: object) -> str:
        raw = cls._normalize_text(value)
        canonical = CANONICAL_PORTAL_MAP.get(raw)
        if canonical:
            return canonical
        return raw.replace(" ", "_")

    @staticmethod
    def _build_folder_name(proveedor: str, sede_subportal: str) -> str:
        if sede_subportal:
            return f"{proveedor}_{sede_subportal}"
        return proveedor
